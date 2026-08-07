"""Tests for po_materials/estimate_parse.py (+ estimate_ocr) — the Tier-1
deterministic extraction tier (ADR-0004 E4).

Fixtures are SYNTHETIC text replicas of the corpus layouts (Platt with the
M-divisor case, OnPoint SOV headers, Terratech clean table) — NO real vendor
bytes in the repo. RED musts covered:
  * M-divisor math: 2,500 @ $1,098.90/M = $2,747.25 EXACTLY (274725 cents);
  * tampered extended → math_ok False + flags (delete the math gate → fail);
  * OnPoint SOV/SOC header rows produce ZERO $0 lines (they become sections);
  * to_cents('$993.12429') is deterministic (Decimal ROUND_HALF_UP quantize);
  * check_math NEVER raises, even on NaN garbage.

Run with: pytest -q tests/test_estimate_parse.py
"""
from __future__ import annotations

import io
import json
import sys
import types
from pathlib import Path

import pytest

from po_materials import estimate_ocr, estimate_parse, estimate_preview, estimate_sandbox
from po_materials.estimate_parse import (
    ExtractionResult,
    LineItem,
    ParsedPdf,
    check_math,
    family_key,
    load_vendor_templates,
    parse_generic_table,
    parse_native,
    parse_with_template,
    to_cents,
)

# ---- Synthetic Platt replica (the template fixture) --------------------------------

PLATT_TEXT = """\
PLATT ELECTRIC SUPPLY
Quote #4471205
Updated On: 06/12/2026

PV Source Circuits (String Wire)
1 PVW10BLK | Item# 55501 PV WIRE 10AWG 2000V BLACK 2,500 $1,098.90 M $2,747.25
MFR: SOUTHWIRE
2 CQD230 | Item# 55502 SQ D CIRCUIT BREAKER 30A 4 $45.10 EA $180.40
* Expected stock arrival 07/01
Inverters & BOS
3 TAPE33 | Item# 55503 ELECTRICAL TAPE 3/4IN 12 $0.45 EA $5.40

SUB TOTAL: $2,933.05
TAX: $205.31
TOTAL: $3,138.36
"""


def _platt_parsed(text: str = PLATT_TEXT) -> ParsedPdf:
    return ParsedPdf(
        pages_text=[text], words=[[]], tables=[[]],
        chars_per_page=[len(text)], is_scanned=False,
    )


def _platt_template() -> estimate_parse.VendorTemplate:
    templates = load_vendor_templates()
    by_name = {t.name: t for t in templates}
    assert "platt" in by_name, "shipped platt.yaml failed to load"
    return by_name["platt"]


# ---- to_cents / family_key ----------------------------------------------------------


def test_to_cents_thousands_commas():
    assert to_cents("1,098.90") == 109890
    assert to_cents("$22,500.00") == 2250000


def test_red_to_cents_overprecise_is_deterministic():
    """'$993.12429' quantizes to cents (ROUND_HALF_UP) — same answer every time."""
    assert to_cents("$993.12429") == 99312
    assert to_cents("$993.12429") == to_cents("$993.12429")
    assert to_cents("$993.125") == 99313  # half rounds UP, not banker's
    assert to_cents("1.005") == 101


def test_to_cents_plain_numbers_and_negatives():
    assert to_cents(45.10) == 4510
    assert to_cents(0) == 0
    assert to_cents("(12.50)") == -1250


@pytest.mark.parametrize(
    "bad", ["", "N/A", "$", "12.3.4", "1/2", float("nan"), None, True, ["1.00"]]
)
def test_to_cents_unparseable_returns_none(bad):
    """The ladder contract (`to_cents(value) -> int | None`): a bad cell degrades
    the field to absent — it never raises into the daemon."""
    assert to_cents(bad) is None


def test_family_key_normalizes_vendor_and_quote():
    assert (
        family_key("Platt Electric Supply", "4471205", "ab" * 32)
        == family_key("PLATT  Electric-Supply,", "4471205", "cd" * 32)
        == "plattelectricsupply|4471205"
    )


def test_family_key_numberless_falls_back_to_sha():
    sha = "AB" * 32
    assert family_key("Brimfield", None, sha) == f"brimfield|{'ab' * 32}"
    assert family_key("Brimfield", "  ", sha).endswith("ab" * 32)


# ---- parse_native (sandbox seam) ----------------------------------------------------


def _mock_sandbox(monkeypatch, payload):
    calls: list[tuple] = []

    def fake(fn_name, data, *, timeout_s, rlimit_bytes=0, args=()):
        calls.append((fn_name, args))
        if payload is None:
            return None
        return json.dumps(payload).encode()

    monkeypatch.setattr(estimate_parse.estimate_sandbox, "run_sandboxed", fake)
    return calls


def test_parse_native_shapes_child_payload(monkeypatch):
    calls = _mock_sandbox(monkeypatch, {
        "pages": ["hello world"],
        "chars_per_page": [500],
        "words": [[{"text": "hello", "x0": 1.0, "x1": 2.0, "top": 3.0, "bottom": 4.0}]],
        "tables": [[[["a", "b"]]]],
    })
    parsed = parse_native(b"%PDF-1.4 fake")
    assert parsed is not None
    assert parsed.pages_text == ["hello world"]
    assert parsed.words[0][0]["text"] == "hello"
    assert parsed.tables[0][0] == [["a", "b"]]
    assert parsed.is_scanned is False
    assert calls[0][0] == "parse_native"


def test_parse_native_is_scanned_on_near_zero_chars(monkeypatch):
    _mock_sandbox(monkeypatch, {
        "pages": ["", " "], "chars_per_page": [0, 3], "words": [[], []], "tables": [[], []],
    })
    parsed = parse_native(b"scanned")
    assert parsed is not None
    assert parsed.is_scanned is True


def test_parse_native_sandbox_failure_degrades_to_none(monkeypatch):
    _mock_sandbox(monkeypatch, None)
    assert parse_native(b"hostile") is None


def test_parse_native_malformed_child_output_degrades(monkeypatch):
    monkeypatch.setattr(
        estimate_parse.estimate_sandbox, "run_sandboxed",
        lambda *a, **k: b"not json",
    )
    assert parse_native(b"x") is None
    monkeypatch.setattr(
        estimate_parse.estimate_sandbox, "run_sandboxed",
        lambda *a, **k: json.dumps({"pages": "nope"}).encode(),
    )
    assert parse_native(b"x") is None


def test_parse_native_live_sandbox_roundtrip():
    """LIVE child spawn (no mock): a reportlab-built native PDF parses through the
    real `python -m po_materials.estimate_sandbox parse_native` child."""
    from reportlab.lib.pagesizes import letter
    from reportlab.pdfgen import canvas

    buf = io.BytesIO()
    c = canvas.Canvas(buf, pagesize=letter)
    c.drawString(72, 720, "Quote #12345 NATIVE TEXT LAYER for the parse ladder")
    c.showPage()
    c.save()
    parsed = parse_native(buf.getvalue())
    assert parsed is not None
    assert "NATIVE TEXT LAYER" in parsed.pages_text[0]
    assert parsed.is_scanned is False
    assert parsed.chars_per_page[0] > estimate_parse.SCANNED_CHARS_PER_PAGE


# ---- Vendor templates ---------------------------------------------------------------


def test_load_vendor_templates_ships_platt():
    tpl = _platt_template()
    assert tpl.vendor_name == "Platt"
    assert tpl.uom_divisors == {"M": 1000}
    assert tpl.doc_type == "quote"


def test_load_vendor_templates_skips_malformed_loudly(tmp_path: Path, caplog):
    (tmp_path / "good.yaml").write_text(
        "name: good\nvendor_name: Good Co\nmatch: ['GOOD']\n"
        "lines:\n  pattern: '(?P<qty>\\d+) x (?P<description>.+) @ \\$(?P<unit_price>[\\d.]+)'\n",
        encoding="utf-8",
    )
    (tmp_path / "bad.yaml").write_text("just a string, not a template", encoding="utf-8")
    (tmp_path / "badre.yaml").write_text(
        "name: badre\nvendor_name: X\nmatch: ['[unclosed']\n", encoding="utf-8"
    )
    with caplog.at_level("WARNING"):
        templates = load_vendor_templates(tmp_path)
    assert [t.name for t in templates] == ["good"]
    assert sum(
        "skipping malformed vendor template" in r.getMessage() for r in caplog.records
    ) == 2


def test_load_vendor_templates_absent_dir_is_empty():
    assert load_vendor_templates(Path("/nonexistent/estimate_templates")) == []


def test_template_rejects_refused_doc_types(tmp_path: Path):
    with pytest.raises(ValueError):
        estimate_parse._template_from_mapping(
            {"name": "x", "vendor_name": "X", "doc_type": "invoice", "match": ["X"]},
            source_path="t",
        )


# ---- Platt template parse (incl. the M-divisor RED must) ----------------------------


def test_platt_parse_extracts_lines_sections_and_header():
    result = parse_with_template(_platt_parsed(), _platt_template())
    assert result is not None
    assert result.tier == "tier1_template"
    assert result.vendor_name == "Platt"
    assert result.quote_number == "4471205"
    assert result.quote_date == "2026-06-12"
    assert len(result.line_items) == 3
    wire, breaker, tape = result.line_items
    assert wire.section == "PV Source Circuits (String Wire)"
    assert breaker.section == "PV Source Circuits (String Wire)"
    assert tape.section == "Inverters & BOS"
    assert wire.part_number == "PVW10BLK"
    assert breaker.qty == 4 and breaker.unit == "EA"
    assert result.subtotal_cents == 293305
    assert result.tax_cents == 20531
    assert result.grand_total_cents == 313836


def test_red_platt_m_divisor_math_exact():
    """2,500 FT of wire at $1,098.90 per THOUSAND (UOM 'M') extends to $2,747.25
    EXACTLY — encode the divisor wrong (or drop it) and this fails."""
    result = parse_with_template(_platt_parsed(), _platt_template())
    assert result is not None
    wire = result.line_items[0]
    assert wire.qty == 2500.0
    assert wire.unit == "M"
    assert wire.unit_cost_cents == 109890
    assert wire.extended_cents == 274725
    assert wire.math_ok is True
    assert result.math_ok is True
    assert result.math_flags == []
    assert result.needs_review is False


def test_red_platt_tampered_extended_trips_math_gate():
    tampered = PLATT_TEXT.replace("$2,747.25", "$9,999.99")
    # (the doc-level Σextended check also fires — the assertion pins the LINE flag)
    result = parse_with_template(_platt_parsed(tampered), _platt_template())
    assert result is not None
    assert result.math_ok is False
    assert result.line_items[0].math_ok is False
    assert any("qty×unit_cost != extended" in f for f in result.math_flags)
    assert result.needs_review is True


def test_platt_stock_notes_and_mfr_lines_never_become_lines():
    result = parse_with_template(_platt_parsed(), _platt_template())
    assert result is not None
    joined = " ".join(li.description for li in result.line_items)
    assert "SOUTHWIRE" not in joined
    assert "Expected stock" not in joined
    assert all(li.extended_cents != 0 for li in result.line_items)


def test_template_not_matching_returns_none():
    other = _platt_parsed("TERRATECH LLC\nProposal for services\n")
    assert parse_with_template(other, _platt_template()) is None


def test_template_match_with_zero_lines_returns_none():
    header_only = _platt_parsed("PLATT ELECTRIC SUPPLY\nQuote #999\nno line rows here\n")
    assert parse_with_template(header_only, _platt_template()) is None


# ---- check_math ---------------------------------------------------------------------


def _result(lines, **totals) -> ExtractionResult:
    return ExtractionResult(
        doc_type="quote", confidence=1.0, line_items=lines, **totals
    )


def test_check_math_doc_level_subtotal_mismatch_flagged():
    lines = [LineItem(description="a", qty=2, unit_cost_cents=100, extended_cents=200)]
    result = check_math(_result(lines, subtotal_cents=999))
    assert result.math_ok is False
    assert any("subtotal" in f for f in result.math_flags)


def test_check_math_grand_total_composition():
    lines = [LineItem(description="a", qty=1, unit_cost_cents=1000, extended_cents=1000)]
    ok = check_math(
        _result(lines, subtotal_cents=1000, tax_cents=70, freight_cents=30,
                grand_total_cents=1100)
    )
    assert ok.math_ok is True
    bad = check_math(_result(lines, subtotal_cents=1000, grand_total_cents=1100))
    assert bad.math_ok is False


def test_check_math_skips_absent_operands():
    lines = [LineItem(description="lump sum not-to-exceed")]  # no qty/price at all
    result = check_math(_result(lines, not_to_exceed_cap_cents=5_000_000))
    assert result.math_ok is True
    assert result.math_flags == []


def test_red_check_math_never_raises_on_garbage():
    lines = [
        LineItem(description="nan bomb", qty=float("nan"), unit_cost_cents=100,
                 extended_cents=100)
    ]
    result = check_math(_result(lines))  # must not raise
    assert result.math_ok is False
    assert result.math_flags


def test_check_math_respects_caller_divisors():
    lines = [LineItem(description="per-hundred", qty=200, unit="C",
                      unit_cost_cents=5000, extended_cents=10000)]
    assert check_math(_result(lines), uom_divisors={"C": 100}).math_ok is True
    assert check_math(_result(lines)).math_ok is False  # without the divisor: 200×5000


# ---- Generic table tier (OnPoint / Terratech replicas) ------------------------------

_ONPOINT_TABLE = [
    ["Item", "Description", "Qty", "UOM", "Unit Price", "Total"],
    ["", "SOLAR PV SYSTEM - SCHEDULE OF VALUES", "", "", "", ""],
    ["1", "Rooftop PV module install", "120", "EA", "$85.00", "$10,200.00"],
    ["", "ELECTRICAL - SOC", "", "", "", ""],
    ["2", "Conduit run", "300", "FT", "$12.50", "$3,750.00"],
]


def _tabled(tables, text="OnPoint Construction\nQuote # OP-889\n") -> ParsedPdf:
    return ParsedPdf(
        pages_text=[text], words=[[]], tables=[tables],
        chars_per_page=[len(text)], is_scanned=False,
    )


def test_red_onpoint_sov_header_rows_produce_zero_dollar_zero_lines():
    """SOV/SOC section-band rows carry no qty/unit price — they become SECTION
    labels, never $0 line items (the corpus rule; delete the guard → this fails)."""
    result = parse_generic_table(_tabled([_ONPOINT_TABLE]))
    assert result is not None
    assert len(result.line_items) == 2
    assert all(li.qty not in (0, None) for li in result.line_items)
    assert all(li.unit_cost_cents not in (0, None) for li in result.line_items)
    assert all((li.extended_cents or 1) != 0 for li in result.line_items)
    assert result.line_items[0].section == "SOLAR PV SYSTEM - SCHEDULE OF VALUES"
    assert result.line_items[1].section == "ELECTRICAL - SOC"


def test_generic_table_terratech_clean_parse_with_totals():
    text = (
        "Terratech Inc\nQuotation Number: TT-2201\n"
        "Subtotal: $13,950.00\nSales Tax: $976.50\nTotal: $14,926.50\n"
    )
    result = parse_generic_table(_tabled([_ONPOINT_TABLE], text=text))
    assert result is not None
    assert result.tier == "tier1_generic"
    assert result.vendor_name == "Terratech Inc"
    assert result.subtotal_cents == 1395000
    assert result.tax_cents == 97650
    assert result.grand_total_cents == 1492650
    # Σextended 1020000 + 375000 == subtotal → math gate green
    assert result.math_ok is True
    assert result.needs_review is False


def test_generic_table_none_when_no_table_infers():
    narrative = _tabled([], text="We propose to perform the work for a not-to-exceed sum")
    assert parse_generic_table(narrative) is None
    headerless = _tabled([[["just", "words"], ["no", "header"]]])
    assert parse_generic_table(headerless) is None


def test_generic_table_math_tamper_flags_needs_review():
    tampered = [row[:] for row in _ONPOINT_TABLE]
    tampered[2] = ["1", "Rooftop PV module install", "120", "EA", "$85.00", "$11,111.11"]
    result = parse_generic_table(_tabled([tampered]))
    assert result is not None
    assert result.math_ok is False
    assert result.needs_review is True


# ---- estimate_ocr (Vision seam) -----------------------------------------------------


def _fake_ocrmac(monkeypatch, annotations_per_call, raise_on_call=False):
    fake_pkg = types.ModuleType("ocrmac")
    fake_mod = types.ModuleType("ocrmac.ocrmac")

    class OCR:
        def __init__(self, image, recognition_level="fast"):
            assert recognition_level == "accurate"

        def recognize(self):
            if raise_on_call:
                raise RuntimeError("vision exploded")
            return annotations_per_call

    setattr(fake_mod, "OCR", OCR)  # noqa: B010 — dynamic module attr by design
    setattr(fake_pkg, "ocrmac", fake_mod)  # noqa: B010
    monkeypatch.setitem(sys.modules, "ocrmac", fake_pkg)
    monkeypatch.setitem(sys.modules, "ocrmac.ocrmac", fake_mod)


def _tiny_png() -> bytes:
    from PIL import Image

    buf = io.BytesIO()
    Image.new("RGB", (4, 4), "white").save(buf, format="PNG")
    return buf.getvalue()


def test_ocr_pages_happy_path(monkeypatch):
    _fake_ocrmac(monkeypatch, [("Hello", 0.99, (0, 0, 1, 1)), ("World", 0.98, (0, 0, 1, 1))])
    monkeypatch.setattr(
        estimate_ocr.estimate_preview, "render_page_pngs",
        lambda data, *, max_pages: [_tiny_png(), _tiny_png()],
    )
    pages = estimate_ocr.ocr_pages(b"scanned pdf", max_pages=4)
    assert pages == ["Hello\nWorld", "Hello\nWorld"]


def test_ocr_pages_render_failure_degrades_to_empty(monkeypatch):
    _fake_ocrmac(monkeypatch, [("x", 0.9, (0, 0, 1, 1))])
    monkeypatch.setattr(
        estimate_ocr.estimate_preview, "render_page_pngs",
        lambda data, *, max_pages: [],
    )
    assert estimate_ocr.ocr_pages(b"bad", max_pages=4) == []


def test_ocr_pages_per_page_failure_degrades_to_blank(monkeypatch):
    _fake_ocrmac(monkeypatch, [], raise_on_call=True)
    monkeypatch.setattr(
        estimate_ocr.estimate_preview, "render_page_pngs",
        lambda data, *, max_pages: [_tiny_png()],
    )
    assert estimate_ocr.ocr_pages(b"pdf", max_pages=4) == [""]


def test_ocr_pages_missing_ocrmac_degrades(monkeypatch):
    monkeypatch.setitem(sys.modules, "ocrmac", None)  # import → ImportError
    monkeypatch.setattr(
        estimate_ocr.estimate_preview, "render_page_pngs",
        lambda data, *, max_pages: (_ for _ in ()).throw(AssertionError("render before import check")),
    )
    assert estimate_ocr.ocr_pages(b"pdf", max_pages=4) == []


# ---- parse_estimate + to_worker_payload (the estimate_poll ladder seam) -------------


def test_parse_estimate_platt_pages_yield_worker_payload():
    """The documented ladder contract: parse_estimate(pages, *, filename) → the
    Worker extraction-body dict (daemon stamps tier)."""
    payload = estimate_parse.parse_estimate([PLATT_TEXT], filename="platt (2).pdf")
    assert payload is not None
    assert payload["schema_version"] == "1.0.0"
    assert payload["doc_type"] == "quote"
    assert payload["vendor_name"] == "Platt"
    assert payload["quote_number"] == "4471205"
    assert payload["math_ok"] == 1
    assert payload["subtotal_cents"] == 293305
    assert isinstance(payload["payload_json"], str) and len(payload["payload_json"]) >= 2
    assert [ln["position"] for ln in payload["lines"]] == [1, 2, 3]
    assert all(ln["math_ok"] == 1 for ln in payload["lines"])
    assert payload["lines"][0]["extended_cents"] == 274725  # the M-divisor line
    assert "tier" not in payload  # the daemon stamps the tier


def test_parse_estimate_unmatched_text_returns_none():
    assert estimate_parse.parse_estimate(["nothing recognizable here"], filename="x.pdf") is None
    assert estimate_parse.parse_estimate([], filename="x.pdf") is None
    assert estimate_parse.parse_estimate(["", "  "], filename="x.pdf") is None


def test_parse_estimate_with_data_reaches_generic_table_tier(monkeypatch):
    """Passing raw bytes lets parse_estimate run the sandbox parse and reach the
    generic-table tier (text alone carries no table geometry)."""
    monkeypatch.setattr(
        estimate_parse.estimate_sandbox, "run_sandboxed",
        lambda *a, **k: json.dumps({
            "pages": ["OnPoint Construction\nQuote # OP-889\n"],
            "chars_per_page": [500],
            "words": [[]],
            "tables": [[_ONPOINT_TABLE]],
        }).encode(),
    )
    payload = estimate_parse.parse_estimate(
        ["OnPoint Construction\nQuote # OP-889\n"], filename="op.pdf", data=b"%PDF-fake"
    )
    assert payload is not None
    assert len(payload["lines"]) == 2
    assert payload["lines"][0]["section"] == "SOLAR PV SYSTEM - SCHEDULE OF VALUES"


def test_to_worker_payload_tamper_surfaces_flags():
    tampered = PLATT_TEXT.replace("$2,747.25", "$9,999.99")
    result = parse_with_template(_platt_parsed(tampered), _platt_template())
    assert result is not None
    payload = estimate_parse.to_worker_payload(result)
    assert payload["math_ok"] == 0
    assert payload["lines"][0]["math_ok"] == 0
    assert payload["anomalies"] and "qty×unit_cost != extended" in payload["anomalies"]
    assert '"needs_review":true' in payload["payload_json"]


def test_to_worker_payload_caps_lines_at_worker_bound():
    lines = [
        LineItem(description=f"item {i}", qty=1, unit_cost_cents=100, extended_cents=100)
        for i in range(600)
    ]
    payload = estimate_parse.to_worker_payload(_result(lines))
    assert len(payload["lines"]) == estimate_parse.MAX_WORKER_LINES
    assert payload["lines"][-1]["position"] == estimate_parse.MAX_WORKER_LINES


# ---- Sandbox registration -----------------------------------------------------------


def test_parse_native_is_a_registered_sandbox_fn():
    """The E4 child fn is allowlisted in the sandbox dispatch (additive edit)."""
    assert "parse_native" in estimate_sandbox._ALLOWED_FNS
    assert estimate_sandbox.PARSE_TIMEOUT_S > 0


def test_preview_module_untouched_contract():
    """estimate_ocr leans on estimate_preview's laundered PNGs — pin the seam."""
    assert callable(estimate_preview.render_page_pngs)


# ---- Tier-1 xlsx (vendor spreadsheets) -----------------------------------------------
#
# These drive the REAL sandbox child (a subprocess), like the parse_native live round-trip
# above — a mocked child would prove nothing about the openpyxl behaviour, which IS the tier.


def _vendor_book(
    *, subtotal=3945.00, tax=None, grand=3945.00, merged_header=False, sheets=1
) -> bytes:
    """A VENDOR quote workbook — deliberately not ITS's own form (no _ITS_META sheet),
    so Tier 0 declines it and it reaches the xlsx tier."""
    import openpyxl

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Quote"
    ws.append(["ACME Electric Supply"])
    ws.append(["Quote #: Q-88213"])
    ws.append([None])
    ws.append(["Part", "Description", "Qty", "UOM", "Unit Price", "Ext Price"])
    ws.append(["THHN-10", "10 AWG THHN Copper", 5000, "FT", 0.42, 2100.00])
    ws.append(["PVC-2", '2" PVC Conduit', 300, "EA", 6.15, 1845.00])
    if merged_header:
        # A VERTICAL merge over the two data rows — the common "one category cell spanning
        # its member rows" layout. iter_rows leaves the second row's cell None.
        ws.merge_cells(start_row=5, start_column=1, end_row=6, end_column=1)
    if subtotal is not None:
        ws.append([None, None, None, None, "Subtotal", subtotal])
    if tax is not None:
        ws.append([None, None, None, None, "Tax", tax])
    if grand is not None:
        ws.append([None, None, None, None, "Grand Total", grand])
    for extra in range(sheets - 1):
        wb.create_sheet(f"Terms{extra}").append(["Boilerplate terms and conditions"])
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def test_parse_xlsx_estimate_extracts_lines_and_cell_totals():
    result = estimate_parse.parse_xlsx_estimate(_vendor_book())
    assert result is not None
    assert result.tier == "tier1_xlsx"
    assert result.vendor_name == "ACME Electric Supply"
    assert result.quote_number == "Q-88213"
    assert [li.part_number for li in result.line_items] == ["THHN-10", "PVC-2"]
    # Integer cents, exact: 5000 FT @ $0.42 = $2,100.00
    assert result.line_items[0].unit_cost_cents == 42
    assert result.line_items[0].extended_cents == 210000
    assert result.subtotal_cents == 394500  # read from the CELL, not regexed
    assert result.math_ok is True
    assert result.math_flags == []


def test_round_totals_are_read_from_cells_not_regexed_from_text():
    """THE control this tier exists for.

    openpyxl returns the STORED value, so a ROUND money amount stringifies with ONE
    decimal (`str(4000.00) == "4000.0"`) while `_GENERIC_TOTALS` demands two. Regexing the
    flattened grid therefore drops round totals — and `check_math` SKIPS comparisons with
    absent operands, so the cross-check silently evaporates.
    """
    data = _vendor_book(subtotal=3945.00, grand=3945.00)
    parsed, raw = estimate_parse.parse_xlsx(data)

    # The text-regex path cannot see the round subtotal at all...
    assert estimate_parse.parse_generic_table(parsed).subtotal_cents is None
    # ...the cell-grid reader can.
    assert estimate_parse._xlsx_totals_from_grid(raw)["subtotal_cents"] == 394500


def test_a_wrong_round_subtotal_is_caught_not_silently_passed():
    """The failure the control prevents: lines sum to $3,945.00, the book claims $3,900.00.

    Asserts the CATCH, so the control cannot rot into a no-op without RED-lighting here.
    """
    data = _vendor_book(subtotal=3900.00, grand=3900.00)

    parsed, _raw = estimate_parse.parse_xlsx(data)
    missed = estimate_parse.check_math(estimate_parse.parse_generic_table(parsed))
    assert missed.math_ok is True  # regex path: silently unverified, error invisible

    caught = estimate_parse.parse_xlsx_estimate(data)
    assert caught.math_ok is False
    assert any("subtotal" in f for f in caught.math_flags)


def test_vertically_merged_cell_fills_down_and_never_across():
    """A merge spanning rows fills DOWN (the value genuinely applies to each row); a merge
    must never fill ACROSS, which would copy one column's value into a sibling column that
    has none. Filling across was measured to overwrite the "Qty" header with "Description",
    after which `_infer_columns` claimed nothing and the sheet yielded ZERO lines."""
    result = estimate_parse.parse_xlsx_estimate(_vendor_book(merged_header=True))
    assert result is not None
    assert len(result.line_items) == 2
    # The merged part-number cell propagated down to the second row...
    assert result.line_items[1].part_number == "THHN-10"
    # ...while every OTHER column of that row kept its own distinct value.
    assert result.line_items[1].qty == 300
    assert result.line_items[1].extended_cents == 184500


def test_extra_sheets_do_not_break_the_parse():
    """Vendor books carry cover / T&C sheets; the line sheet must still parse."""
    result = estimate_parse.parse_xlsx_estimate(_vendor_book(sheets=3))
    assert result is not None
    assert len(result.line_items) == 2


@pytest.mark.parametrize(
    "payload",
    [b"", b"not a zip at all", b"PK\x03\x04" + b"\x00" * 64],
    ids=["empty", "not-a-zip", "zip-magic-garbage"],
)
def test_hostile_bytes_degrade_to_none_and_never_raise(payload):
    """A workbook openpyxl cannot open must DEGRADE, never raise — the ladder contract is
    that a tier failure falls through to needs_review."""
    assert estimate_parse.parse_xlsx(payload) is None
    assert estimate_parse.parse_xlsx_estimate(payload) is None


def test_parse_xlsx_never_reports_scanned():
    """A workbook has no text layer to be missing; the OCR concept does not apply."""
    parsed, _raw = estimate_parse.parse_xlsx(_vendor_book())
    assert parsed.is_scanned is False


def test_xlsx_child_fn_is_sandbox_allowlisted():
    """The hostile parse runs in the killable child, not in the daemon. (Contrast Tier 0,
    which parses in-process — defensible only for ITS's own HMAC-verified form.)"""
    assert "parse_xlsx_grid" in estimate_sandbox._ALLOWED_FNS
    assert estimate_sandbox.XLSX_TIMEOUT_S > 0


# ---- Header keyword matching: not-inside-a-larger-word --------------------------------


def test_red_um_keyword_must_not_match_inside_part_number():
    """RED must. `_COLUMN_CONCEPTS` gives `unit` the short keyword "um", and plain
    SUBSTRING matching found it inside "Part N-um-ber" — so `unit` claimed the
    part-number column and `part_number`, checked later and barred from re-claiming a
    taken cell, was dropped entirely. Revert `_kw_matches` to `kw in cell` and this fails.

    This is the REAL RFQ quote-form header (`quote_form.render_quote_form`), which is why
    synthetic fixtures using a bare "Part" header never surfaced it.
    """
    header = ["#", "Part Number", "Description", "Qty", "Unit", "Unit Price", "Extended"]
    cols = estimate_parse._infer_columns(header)
    assert cols is not None
    assert header[cols["unit"]] == "Unit"
    assert header[cols["part_number"]] == "Part Number"


def test_kw_matching_keeps_legitimate_substring_labels():
    """The guard must not over-tighten: richer keywords are listed FIRST and still win,
    so "Extended"/"Description" keep matching even though "ext"/"desc" would not."""
    header = ["Item", "Description", "Qty", "UOM", "Unit Price", "Extended"]
    cols = estimate_parse._infer_columns(header)
    assert cols is not None
    assert header[cols["extended"]] == "Extended"
    assert header[cols["description"]] == "Description"
    assert header[cols["unit"]] == "UOM"


# ---- The REAL RFQ quote form through the xlsx tier ------------------------------------


def _filled_rfq_quote_form(secret: bytes = b"test-quote-form-secret") -> bytes:
    """Render ITS's OWN fillable RFQ quote form and fill it the way a vendor would.

    This is the artifact `rfq_poll` sends to every vendor, so it is the single most likely
    spreadsheet to arrive back — worth pinning against the real renderer rather than a
    hand-written fixture.
    """
    import openpyxl

    from po_materials import quote_form

    lines = [
        {"part_number": "PVW10BLK", "description": "PV WIRE 10AWG 2000V BLACK",
         "qty": 2500, "unit": "M"},
        {"part_number": "CQD230", "description": "SQ D CIRCUIT BREAKER 30A",
         "qty": 4, "unit": "EA"},
    ]
    blank = quote_form.render_quote_form(
        "RFQ-2201", "platt", "Kiwi Solar", lines, secret=secret
    )
    wb = openpyxl.load_workbook(io.BytesIO(blank))
    ws = wb[quote_form.SHEET_FORM]
    header = [c.value for c in ws[quote_form.TABLE_HEADER_ROW]]
    up = next(i for i, h in enumerate(header, 1) if h and "unit price" in str(h).lower())
    ext = next(i for i, h in enumerate(header, 1) if h and "extended" in str(h).lower())
    for off, (price, extended) in enumerate([(1098.90, 2747.25), (45.10, 180.40)]):
        ws.cell(row=quote_form.FIRST_LINE_ROW + off, column=up, value=price)
        ws.cell(row=quote_form.FIRST_LINE_ROW + off, column=ext, value=extended)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def test_real_rfq_quote_form_parses_through_the_xlsx_tier():
    """The operator's acceptance bar: the tier must handle the RFQ Excel sheets.

    Tier 0 claims a form whose HMAC verifies; this is the FALL-THROUGH path (e.g. a vendor
    who re-saved the workbook and stripped the token). It must still extract cleanly rather
    than dumping the quote on someone to re-type.
    """
    result = estimate_parse.parse_xlsx_estimate(_filled_rfq_quote_form())
    assert result is not None
    assert result.tier == "tier1_xlsx"
    assert [li.part_number for li in result.line_items] == ["PVW10BLK", "CQD230"]
    assert [li.unit for li in result.line_items] == ["M", "EA"]
    assert result.math_ok is True


def test_real_rfq_quote_form_m_divisor_math_exact():
    """2,500 of wire at $1,098.90 per THOUSAND ("M") extends to $2,747.25 EXACTLY.

    `check_math` seeds `DEFAULT_UOM_DIVISORS = {"M": 1000}`, so per-thousand distributor
    pricing reconciles without the caller passing divisors — drop that default and this
    line math-flags.
    """
    result = estimate_parse.parse_xlsx_estimate(_filled_rfq_quote_form())
    wire = result.line_items[0]
    assert wire.qty == 2500.0
    assert wire.unit == "M"
    assert wire.unit_cost_cents == 109890
    assert wire.extended_cents == 274725
    assert wire.math_ok is True


def test_tier0_still_claims_a_verified_quote_form():
    """Regression fence: the xlsx tier must not change Tier-0 behaviour for a form whose
    token verifies — Tier 0 is checked first and stays authoritative."""
    from po_materials import quote_form

    secret = b"test-quote-form-secret"
    parsed = quote_form.parse_quote_form(_filled_rfq_quote_form(secret), secret=secret)
    assert parsed is not None
    assert parsed.verified is True
    assert len(parsed.lines) == 2
