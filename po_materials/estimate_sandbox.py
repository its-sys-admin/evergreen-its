"""Killable-subprocess isolation for hostile-document parsing (ADR-0004, red-team #5).

Purpose
-------
Every hostile-input parse stage of the vendor-estimate importer — pdfplumber text
extraction (`estimate_classify.extract_pages_text`) and the Quartz PDF→PNG preview
render (`estimate_preview.render_page_pngs`) — runs INSIDE a killable child process
with an `RLIMIT_AS` address-space cap and a wall-clock timeout. A wedged / OOM /
crashing parse of an attacker-crafted PDF is REAPED by the parent; the stage returns
None/[] and the document DEGRADES (no previews / doc_type 'other' → needs_review) —
the `estimate_poll` daemon NEVER dies from a hostile document.

Protocol
--------
`run_sandboxed(fn_name, data, ...)` re-invokes `python -m po_materials.estimate_sandbox
<fn_name> [args...]`, writes `data` to the child's stdin, and returns the child's raw
stdout bytes (a JSON document per the fn contract below) — or None on timeout / kill /
crash / nonzero exit / empty output. The parse libraries (pdfplumber, Quartz) are
imported lazily INSIDE the child only — the daemon process never loads them, so a
parser bug cannot corrupt daemon state even in-process.

Child fn contracts (stdout JSON):
  extract_pages_text [max_pages]  → {"pages": ["page 1 text", ...]}
  render_page_pngs   [max_pages]  → {"pngs": ["<base64 png>", ...]}
  parse_native       [max_pages]  → {"pages": [...], "chars_per_page": [int, ...],
                                     "words": [[{"text","x0","x1","top","bottom"}...]...],
                                     "tables": [[table…]…]}  (E4 Tier-1 parse payload;
                                     per-page word positions bounded by
                                     PARSE_MAX_WORDS_PER_PAGE, tables by
                                     PARSE_MAX_TABLES_PER_PAGE × PARSE_MAX_TABLE_ROWS)
  parse_xlsx_grid    [max_sheets] → {"sheets": ["Sheet1", ...],
                                     "tables": [[["cell", ...], ...], ...],  (STRING cells)
                                     "text":   ["tab/newline-joined grid", ...],
                                     "raw":    [[[native, ...], ...], ...]}  (str|int|float|None)
                                    Tier-1 xlsx payload. `raw` preserves NUMERIC TYPES so the
                                    parent reads doc-level totals off the CELL GRID — regexing
                                    the flattened text misses them (openpyxl yields 4685.0,
                                    one decimal; _GENERIC_TOTALS demands two), and an absent
                                    total makes check_math SKIP the cross-check entirely.
                                    Bounded by XLSX_MAX_SHEETS / _ROWS_PER_SHEET / _COLS_PER_ROW.
  extract_xlsx_rows  [max_sheets]  → {"sheets": [{"name": str, "rows": [[cell, ...], ...]},
                                     ...]}  (PR3b materials-manifest grid extraction;
                                     cells ride as JSON scalars — null/bool/int/float/str
                                     — so the parser's normalize_cell still sees the real
                                     type and a part number read as 7006955 does not become
                                     "7006955.0"; datetimes are str()'d for the parser's ISO
                                     regex. Bounded by MANIFEST_XLSX_MAX_SHEETS ×
                                     MANIFEST_XLSX_MAX_ROWS_PER_SHEET × MANIFEST_XLSX_MAX_COLS_PER_ROW)
  (plus four harmless _test_* fns — spin / bounded-alloc / crash / echo — dispatched
  only by tests/test_estimate_sandbox.py to prove the reap contract on REAL children)

Invariants
----------
* The child attempts `resource.setrlimit(RLIMIT_AS, cap)` via preexec_fn BEFORE
  exec — an allocation bomb hits the rlimit and dies in the child, not the daemon.
  HONESTY (live-probed 2026-07-19 on the exec host): the Darwin kernel REJECTS
  lowering RLIMIT_AS/RLIMIT_DATA (EINVAL), so on macOS the AS cap is best-effort
  and the ENFORCED bounds are `RLIMIT_CPU` (set to the wall-clock budget — kills a
  CPU-spinning parse) + the parent-side `subprocess.run(..., timeout=)` kill. The
  isolation guarantee that matters — a wedged/OOM parse dies in the CHILD, never
  the daemon — holds either way. Every setrlimit is try/except-wrapped: a raising
  preexec_fn would otherwise abort the spawn itself.
  (preexec_fn is documented thread-unsafe; the ITS daemons are single-threaded
  one-shot launchd processes, so the constraint holds by construction.)
* `subprocess.run(..., timeout=)` kills the child on wall-clock overrun.
* NEVER raises on hostile input: every failure mode collapses to None. A None is
  the caller's degrade signal, not an error.
* No AI, no network, no sends — pure local parsing (capability-gated with the
  importer: the estimate lane is AI-free by ADR-0004 decision 1).

Consumers
---------
`po_materials/estimate_classify.py` (text extraction → doc-type classifier) and
`po_materials/estimate_preview.py` (page-preview PNG render), both driven by
`po_materials/estimate_poll.py`.
"""
from __future__ import annotations

import base64
import hashlib
import io
import json
import resource
import subprocess
import sys
from collections.abc import Sequence
from typing import Any, NoReturn

# Address-space cap for a parse child: generous for a legitimate 10 MB estimate,
# fatal for an allocation bomb. 2 GiB.
DEFAULT_RLIMIT_BYTES = 2 * 1024 * 1024 * 1024
# Wall-clock budgets (the caller passes these; constants here so both callers and
# tests share one source).
TEXT_TIMEOUT_S = 60
PREVIEW_TIMEOUT_S = 120
# E4 Tier-1 parse (text + words + tables) does strictly more work than the plain
# text extraction — its own budget, still bounded well under the daemon interval.
PARSE_TIMEOUT_S = 90
# xlsx grid parse (BOTH lanes: Tier-1 vendor books and PR3b manifests, which chose
# the same budget). openpyxl read-only iteration is cheap per cell, but a workbook may
# declare enormous extents (the Deep Lake log claims 1,247 x 92 and holds 57 x 12).
XLSX_TIMEOUT_S = 90
# Sanity cap on child stdout — a preview batch of a dozen page PNGs sits far below
# this; anything larger is a runaway child, treated as a failure.
MAX_CHILD_STDOUT_BYTES = 64 * 1024 * 1024

# Preview render geometry (child-side): target on-screen readable width in pixels,
# hard pixel-area cap per page (matches po_attach_screen.MAX_IMAGE_PIXELS posture).
PREVIEW_TARGET_WIDTH_PX = 1100
PREVIEW_MAX_PIXELS = 24_000_000

# Child-side output bounds for the E4 parse payload — a hostile PDF that fabricates
# millions of words/table rows is truncated in the CHILD, keeping stdout bounded.
PARSE_MAX_WORDS_PER_PAGE = 3000
PARSE_MAX_TABLES_PER_PAGE = 20
PARSE_MAX_TABLE_ROWS = 500

# Tier-1 xlsx-grid parse (evergreen #14). A VENDOR workbook is arbitrary bytes from outside
# the tenant, so it gets the same killable-child treatment as a PDF: openpyxl is a zip + XML
# parser and is no safer than pdfplumber. (Contrast Tier 0, `quote_form.parse_quote_form`,
# which parses openpyxl IN-PROCESS — defensible only because that file is ITS's OWN
# fixed-geometry form, HMAC-verified before parsing. A vendor book earns no such assumption.)
# ADR-0004 decision 5 names openpyxl as a stage that belongs in the subprocess.
XLSX_MAX_SHEETS = 12
XLSX_MAX_ROWS_PER_SHEET = 2000
XLSX_MAX_COLS_PER_ROW = 64

# Materials-manifest grid extraction (PR3b) — SEPARATE constants, deliberately.
#
# The two xlsx lanes were built independently on two repositories and collided on these
# exact names with DIFFERENT values. Collapsing them into one set is not a tidy-up: Python's
# last assignment wins, so one lane would silently inherit the other's bounds — and both
# bounds tests read the constants SYMBOLICALLY, so they stay green at whatever number wins.
# Nothing would have caught it.
#
# The values differ because the risk differs. A vendor book is stranger input and keeps the
# tighter caps above; a manifest is an office upload for a known job, sized against the real
# corpus (the largest sample BOM is well under 2,000 rows; the widest declares 92 columns
# behind 12 real ones). Enforced AS THE GRID IS BUILT, not after — MAX_CHILD_STDOUT_BYTES is
# a parent-side check that only fires once the child has already materialized the payload,
# so a workbook declaring a million rows would exhaust the child before the parent refused it.
MANIFEST_XLSX_MAX_SHEETS = 20
MANIFEST_XLSX_MAX_ROWS_PER_SHEET = 5000
MANIFEST_XLSX_MAX_COLS_PER_ROW = 200
# One hostile cell holding megabytes of text would otherwise ride into D1 and then into the
# validate screen's grid.
MANIFEST_XLSX_MAX_CELL_CHARS = 2000

# TEST-SUPPORT child fns (tests/test_estimate_sandbox.py — the REAL-child-process
# suite proving the reap/rlimit contract without a hostile document). Deliberately
# ungated: each is harmless — local CPU/memory inside a child the parent reaps
# (spin / bounded alloc / crash / echo); nothing in the daemon dispatches them,
# and invoking one by hand just burns a few seconds of local CPU.
_TEST_FNS = ("_test_spin", "_test_alloc", "_test_crash", "_test_echo")
_ALLOWED_FNS = (
    "extract_pages_text",
    "render_page_pngs",
    "parse_native",
    "parse_xlsx_grid",
    "extract_xlsx_rows",
    *_TEST_FNS,
)


def run_sandboxed(
    fn_name: str,
    data: bytes,
    *,
    timeout_s: int,
    rlimit_bytes: int = DEFAULT_RLIMIT_BYTES,
    args: Sequence[str] = (),
) -> bytes | None:
    """Run one child-side parse fn over `data` in an rlimited, timeout-bounded child.

    Returns the child's raw stdout bytes (JSON per the fn contract) or None on ANY
    failure — timeout (child killed), crash/OOM (nonzero exit or signal), unknown
    fn, spawn failure, or empty/oversized output. Never raises on hostile input.
    """
    if fn_name not in _ALLOWED_FNS:
        return None

    def _limit_child() -> None:  # pragma: no cover — runs in the child pre-exec
        # BEST-EFFORT address-space cap: Darwin rejects lowering RLIMIT_AS (EINVAL;
        # see the module docstring's honesty note) — never let a raising preexec_fn
        # abort the spawn itself.
        try:
            resource.setrlimit(resource.RLIMIT_AS, (rlimit_bytes, rlimit_bytes))
        except (ValueError, OSError):
            pass
        # ENFORCED CPU-time cap (works on Darwin): a CPU-spinning parse dies at the
        # wall-clock budget even if the parent were gone; the parent timeout kills
        # sleep/IO wedges.
        try:
            resource.setrlimit(resource.RLIMIT_CPU, (timeout_s, timeout_s + 5))
        except (ValueError, OSError):
            pass

    try:
        proc = subprocess.run(  # noqa: S603 — fixed argv, our own interpreter/module
            [sys.executable, "-m", "po_materials.estimate_sandbox", fn_name, *args],
            input=data,
            capture_output=True,
            timeout=timeout_s,
            preexec_fn=_limit_child,  # noqa: PLW1509 — single-threaded daemon (see docstring)
        )
    except subprocess.TimeoutExpired:
        return None
    except (OSError, ValueError):
        return None
    if proc.returncode != 0:
        return None
    out = proc.stdout
    if not out or len(out) > MAX_CHILD_STDOUT_BYTES:
        return None
    return out


# ---- Child-side workers (run ONLY inside the rlimited child) ------------------------
#
# Hostile-input parsing lives below this line. These functions are invoked via the
# __main__ dispatch; the parse libraries are imported lazily HERE so the daemon
# process never loads them. Any exception escaping a worker exits the child nonzero
# — the parent maps that to None (degrade), never a daemon death.


def _child_extract_pages_text(data: bytes, max_pages: int) -> dict[str, Any]:
    """pdfplumber per-page text extraction (child-side). Empty string per unreadable
    page; a completely unparseable document raises (→ nonzero exit → parent None)."""
    import pdfplumber  # noqa: PLC0415 — lazy child-only import by design

    pages: list[str] = []
    with pdfplumber.open(io.BytesIO(data)) as pdf:
        for page in pdf.pages[:max_pages]:
            try:
                pages.append(page.extract_text() or "")
            except Exception:  # noqa: BLE001 — one bad page degrades to ""
                pages.append("")
    return {"pages": pages}


def _child_parse_native(data: bytes, max_pages: int) -> dict[str, Any]:
    """E4 Tier-1 parse payload (child-side): per-page text + word positions +
    pdfplumber table extraction + text-char counts (the parent's is_scanned input).

    Per-page failures degrade that page ('' / [] / 0); a document pdfplumber cannot
    open raises (→ nonzero exit → parent None, the degrade signal). All list output
    is bounded by the PARSE_MAX_* caps so a hostile PDF cannot balloon stdout."""
    import pdfplumber  # noqa: PLC0415 — lazy child-only import by design

    pages: list[str] = []
    chars_per_page: list[int] = []
    words: list[list[dict[str, Any]]] = []
    tables: list[list[list[list[str]]]] = []
    with pdfplumber.open(io.BytesIO(data)) as pdf:
        for page in pdf.pages[:max_pages]:
            try:
                pages.append(page.extract_text() or "")
            except Exception:  # noqa: BLE001 — one bad page degrades to ""
                pages.append("")
            try:
                chars_per_page.append(len(page.chars))
            except Exception:  # noqa: BLE001
                chars_per_page.append(0)
            try:
                page_words = [
                    {
                        "text": str(w.get("text", "")),
                        "x0": round(float(w.get("x0", 0.0)), 2),
                        "x1": round(float(w.get("x1", 0.0)), 2),
                        "top": round(float(w.get("top", 0.0)), 2),
                        "bottom": round(float(w.get("bottom", 0.0)), 2),
                    }
                    for w in (page.extract_words() or [])[:PARSE_MAX_WORDS_PER_PAGE]
                ]
            except Exception:  # noqa: BLE001
                page_words = []
            words.append(page_words)
            try:
                page_tables = [
                    [
                        ["" if cell is None else str(cell) for cell in row]
                        for row in (table or [])[:PARSE_MAX_TABLE_ROWS]
                    ]
                    for table in (page.extract_tables() or [])[:PARSE_MAX_TABLES_PER_PAGE]
                ]
            except Exception:  # noqa: BLE001
                page_tables = []
            tables.append(page_tables)
    return {
        "pages": pages,
        "chars_per_page": chars_per_page,
        "words": words,
        "tables": tables,
    }


def _child_extract_xlsx_rows(data: bytes, max_sheets: int) -> dict[str, Any]:
    """openpyxl cell-grid extraction for the PR3b materials-manifest importer
    (child-side).

    WHY THIS IS HERE AND NOT IN THE DAEMON. An office-uploaded BOM / shipping log is
    untrusted portal-inbound content (Invariant 2) and openpyxl is a zip+XML parser
    over those bytes. Running it in-process would put a hostile workbook one parser bug
    away from the daemon's own state; here a wedged / OOM / crashing parse is REAPED by
    the parent and the manifest degrades to `refused`. This closes the ADR-0004
    decision-5 gap where the xlsx path was the one hostile-input parse still running
    in-process.

    CELL TYPES RIDE AS JSON SCALARS, DELIBERATELY. `json.dumps` is the only serializer
    and it cannot encode a `datetime`, so the naive fix is to `str()` every cell — but
    that is exactly wrong for this consumer: `manifest_parse.normalize_cell` renders a
    float `7006955.0` as `"7006955"` (a part number) and a str `"7006955.0"` as
    `"7006955.0"` (a part number that matches nothing). So null/bool/int/float/str pass
    through with their type intact and only the non-JSON-encodable values (datetime,
    date, time, Decimal, and anything exotic) are stringified — `datetime` into the
    `"YYYY-MM-DD 00:00:00"` form that `normalize_cell`'s ISO regex already collapses to
    a date.

    `data_only=True` so a formula cell yields its CACHED VALUE rather than the formula
    text; `read_only=True` for the streaming row iterator. Per-sheet failures degrade
    that sheet to an empty row list; a workbook openpyxl cannot open at all RAISES (→
    nonzero exit → parent None, the degrade signal). All output is bounded as it is
    built by the MANIFEST_XLSX_MAX_* caps, so a workbook that DECLARES a million rows cannot
    balloon the child.
    """
    import openpyxl  # noqa: PLC0415 — lazy child-only import by design

    workbook = openpyxl.load_workbook(
        io.BytesIO(data), read_only=True, data_only=True
    )
    sheets: list[dict[str, Any]] = []
    try:
        for worksheet in workbook.worksheets[:max_sheets]:
            rows: list[list[Any]] = []
            try:
                for row in worksheet.iter_rows(values_only=True):
                    if len(rows) >= MANIFEST_XLSX_MAX_ROWS_PER_SHEET:
                        break
                    rows.append([_json_cell(v) for v in row[:MANIFEST_XLSX_MAX_COLS_PER_ROW]])
            except Exception:  # noqa: BLE001 — one bad sheet degrades to [], not fatal
                rows = []
            sheets.append({"name": str(worksheet.title), "rows": rows})
    finally:
        # read_only workbooks hold an open zip handle; close it even on a bad sheet.
        try:
            workbook.close()
        except Exception:  # noqa: BLE001 — best-effort; the child is about to exit
            pass
    return {"sheets": sheets}


def _json_cell(value: Any) -> Any:
    """One openpyxl cell → a JSON-encodable scalar, PRESERVING numeric type.

    See `_child_extract_xlsx_rows` for why `str()`-ing everything would corrupt part
    numbers. Anything json cannot encode natively becomes its `str()` form, bounded to
    MANIFEST_XLSX_MAX_CELL_CHARS so one hostile cell cannot carry megabytes into D1.
    """
    if value is None or isinstance(value, bool | int | float):
        # bool before int is not needed here (both pass through), but floats that are
        # NaN/inf would break json.dumps(allow_nan default) downstream — refuse them.
        if isinstance(value, float) and (value != value or value in (float("inf"), float("-inf"))):
            return ""
        return value
    text = str(value)
    return text[:MANIFEST_XLSX_MAX_CELL_CHARS]


def _child_render_page_pngs(data: bytes, max_pages: int) -> dict[str, Any]:
    """Quartz (CoreGraphics) PDF→PNG page render (child-side).

    CGPDFDocumentCreateWithProvider → per page CGBitmapContext draw →
    CGImageDestination PNG encode. A page that fails to draw/encode is skipped; a
    document Quartz cannot open yields {"pngs": []} (the parent treats an empty
    list as no-previews, a degrade — the disposition screen then forces the
    no-preview path per ADR-0004 decision 3). If the Quartz bridge is unavailable
    at runtime, returns {"pngs": []} gracefully (contract: never a hard dependency).
    """
    try:
        import Quartz  # noqa: PLC0415 — lazy child-only import by design
    except ImportError:
        return {"pngs": []}

    cf_data = Quartz.CFDataCreate(None, data, len(data))
    provider = Quartz.CGDataProviderCreateWithCFData(cf_data)
    doc = Quartz.CGPDFDocumentCreateWithProvider(provider)
    if doc is None:
        return {"pngs": []}
    page_count = min(int(Quartz.CGPDFDocumentGetNumberOfPages(doc)), max_pages)
    pngs: list[str] = []
    for page_no in range(1, page_count + 1):
        try:
            png = _render_one_page(Quartz, doc, page_no)
        except Exception:  # noqa: BLE001 — one bad page is skipped, not fatal
            png = None
        if png:
            pngs.append(base64.b64encode(png).decode("ascii"))
    return {"pngs": pngs}


def _render_one_page(quartz: Any, doc: Any, page_no: int) -> bytes | None:
    """Render one PDF page to PNG bytes via a white-backed RGB bitmap context."""
    page = quartz.CGPDFDocumentGetPage(doc, page_no)
    if page is None:
        return None
    box = quartz.CGPDFPageGetBoxRect(page, quartz.kCGPDFMediaBox)
    width_pts = float(box.size.width)
    height_pts = float(box.size.height)
    if width_pts <= 0 or height_pts <= 0:
        return None
    scale = PREVIEW_TARGET_WIDTH_PX / width_pts
    scale = max(0.1, min(scale, 3.0))
    if width_pts * scale * height_pts * scale > PREVIEW_MAX_PIXELS:
        # Absurd page geometry (a decompression-bomb-shaped MediaBox) — shrink to cap.
        scale = (PREVIEW_MAX_PIXELS / (width_pts * height_pts)) ** 0.5
    width = max(1, int(width_pts * scale))
    height = max(1, int(height_pts * scale))

    color_space = quartz.CGColorSpaceCreateDeviceRGB()
    ctx = quartz.CGBitmapContextCreate(
        None, width, height, 8, 0, color_space, quartz.kCGImageAlphaPremultipliedLast
    )
    if ctx is None:
        return None
    quartz.CGContextSetRGBFillColor(ctx, 1.0, 1.0, 1.0, 1.0)
    quartz.CGContextFillRect(ctx, quartz.CGRectMake(0, 0, width, height))
    quartz.CGContextScaleCTM(ctx, scale, scale)
    quartz.CGContextTranslateCTM(ctx, -float(box.origin.x), -float(box.origin.y))
    quartz.CGContextDrawPDFPage(ctx, page)
    image = quartz.CGBitmapContextCreateImage(ctx)
    if image is None:
        return None
    out_data = quartz.CFDataCreateMutable(None, 0)
    dest = quartz.CGImageDestinationCreateWithData(out_data, "public.png", 1, None)
    if dest is None:
        return None
    quartz.CGImageDestinationAddImage(dest, image, None)
    if not quartz.CGImageDestinationFinalize(dest):
        return None
    return bytes(out_data)


# ---- Test-support child fns (REAL-kill proof; see tests/test_estimate_sandbox.py) --
#
# _test_alloc BOUNDS its own allocation (_TEST_ALLOC_CAP_BYTES) and then spins:
# on a platform that enforces a lowered RLIMIT_AS (Linux) it dies mid-allocation;
# on Darwin (which REJECTS lowering RLIMIT_AS — the module docstring's honesty
# note) the bound keeps host memory safe until the CPU/wall-clock reap kills it.

_TEST_ALLOC_BLOCK_BYTES = 64 * 1024 * 1024
_TEST_ALLOC_CAP_BYTES = 512 * 1024 * 1024


def _child_test_spin() -> NoReturn:  # pragma: no cover — runs in the child, reaped
    """CPU-spin forever — killed by RLIMIT_CPU or the parent wall-clock timeout."""
    while True:
        pass


def _child_test_alloc() -> NoReturn:  # pragma: no cover — runs in the child, reaped
    """Allocate up to the cap (dies to RLIMIT_AS where enforced), then spin for
    the reap — never exits cleanly, so the parent always maps this to None."""
    blocks: list[bytearray] = []
    total = 0
    while total < _TEST_ALLOC_CAP_BYTES:
        blocks.append(bytearray(_TEST_ALLOC_BLOCK_BYTES))  # zero-fill commits pages
        total += _TEST_ALLOC_BLOCK_BYTES
    while True:  # cap reached without an enforceable RLIMIT_AS: await the reap
        pass


def _child_parse_xlsx_grid(data: bytes, max_sheets: int) -> dict[str, Any]:
    """Tier-1 xlsx payload (child-side): every worksheet as a cell grid.

    Emits THREE parallel per-sheet views, because the consumer needs two different things
    and conflating them is the bug this design exists to avoid:

      tables — rows of STRING cells. Feeds `estimate_parse.parse_generic_table` unchanged,
               which is shape-compatible with a worksheet by construction.
      text   — the same grid tab/newline-joined, for vendor-name / quote-number regexes.
      raw    — rows of NATIVE values (str | float | int | None). **This is the load-bearing
               one for doc-level totals.** openpyxl returns the STORED value, and a ROUND
               money amount stringifies with ONE decimal (`str(4000.00) == "4000.0"`) while
               `estimate_parse._GENERIC_TOTALS` requires exactly `\\.\\d{2}`. Regexing the
               flattened text therefore drops round totals — subtotals especially — and
               `check_math` SKIPS comparisons with absent operands, so a book with a WRONG
               subtotal posts as `extracted`, math_ok=True, cross-check never performed.
               The parent reads totals off `raw` and hands the native value straight to
               `to_cents`, which takes floats exactly. See
               `estimate_parse._xlsx_totals_from_grid` for the measured before/after.

    Merged cells are filled FORWARD across their range: `iter_rows` puts the value in the
    top-left and `None` everywhere else, so a merged header band would otherwise blank the
    columns and `_infer_columns` would fail to claim the header row. This needs
    `ws.merged_cells`, which openpyxl does NOT expose under `read_only=True` — hence a
    normal (non-read-only) load, whose memory is bounded by the sandbox RLIMIT + the caps
    above rather than by streaming.

    A workbook openpyxl cannot open raises (→ nonzero exit → parent None, the degrade
    signal). Per-sheet failures degrade that sheet to empty rather than killing the parse.
    """
    import openpyxl  # noqa: PLC0415 — lazy child-only import by design

    wb = openpyxl.load_workbook(io.BytesIO(data), data_only=True, keep_vba=False)
    sheets: list[str] = []
    tables: list[list[list[str]]] = []
    texts: list[str] = []
    raws: list[list[list[Any]]] = []

    for ws in wb.worksheets[:max_sheets]:
        sheets.append(str(ws.title))
        try:
            grid: list[list[Any]] = [
                list(row[:XLSX_MAX_COLS_PER_ROW])
                for row in ws.iter_rows(values_only=True, max_row=XLSX_MAX_ROWS_PER_SHEET)
            ]
            # Merged ranges: fill DOWN the anchor column only — never ACROSS.
            #
            # `iter_rows` puts the value in the top-left and None everywhere else. Filling a
            # merge DOWN is semantically sound (a category/section cell merged over its member
            # rows genuinely applies to each of them). Filling ACROSS is NOT: it copies one
            # column's value into a neighbouring column that has no such value, fabricating
            # data. Measured — merging the "Description" and "Qty" headers and filling across
            # overwrote the Qty header, so `_infer_columns` could not claim the column and the
            # sheet yielded ZERO lines. A merge must never be able to destroy a sibling column.
            merged = getattr(ws, "merged_cells", None)
            for rng in (merged.ranges if merged is not None else []):
                r0, c0, r1 = rng.min_row, rng.min_col, rng.max_row
                col = c0 - 1
                if r0 - 1 >= len(grid) or col >= XLSX_MAX_COLS_PER_ROW:
                    continue
                anchor = grid[r0 - 1][col] if col < len(grid[r0 - 1]) else None
                if anchor is None or r1 <= r0:
                    continue
                for r in range(r0, min(r1, len(grid))):  # rows BELOW the anchor
                    if col < len(grid[r]) and grid[r][col] is None:
                        grid[r][col] = anchor
        except Exception:  # noqa: BLE001 — one bad sheet degrades to empty, never fatal
            grid = []

        # JSON-safe natives only: openpyxl can hand back datetime/Decimal/etc.
        raw_rows: list[list[Any]] = [
            [c if isinstance(c, (str, int, float)) or c is None else str(c) for c in row]
            for row in grid
        ]
        str_rows: list[list[str]] = [
            ["" if c is None else str(c) for c in row] for row in raw_rows
        ]
        raws.append(raw_rows)
        tables.append(str_rows)
        texts.append("\n".join("\t".join(r) for r in str_rows))

    return {"sheets": sheets, "tables": tables, "text": texts, "raw": raws}


def _child_test_echo(data: bytes) -> dict[str, Any]:
    """Happy-path round-trip probe: prove stdin bytes reached the child intact
    and the JSON-on-stdout contract works end-to-end."""
    return {"echo_len": len(data), "echo_sha256": hashlib.sha256(data).hexdigest()}


def _child_main(argv: list[str]) -> int:
    """Child entry: dispatch `<fn_name> [max_pages]`, data on stdin, JSON on stdout."""
    if len(argv) < 1 or argv[0] not in _ALLOWED_FNS:
        return 2
    fn_name = argv[0]
    try:
        max_pages = int(argv[1]) if len(argv) > 1 else 8
    except ValueError:
        return 2
    max_pages = max(1, min(max_pages, 50))
    data = sys.stdin.buffer.read()
    if fn_name == "extract_pages_text":
        result = _child_extract_pages_text(data, max_pages)
    elif fn_name == "parse_native":
        result = _child_parse_native(data, max_pages)
    elif fn_name == "parse_xlsx_grid":
        result = _child_parse_xlsx_grid(data, max_pages)
    elif fn_name == "render_page_pngs":
        result = _child_render_page_pngs(data, max_pages)
    # NOTE the trailing `else:` below is an UNGUARDED fall-through to _child_test_alloc.
    # Adding a name to _ALLOWED_FNS without adding its branch here does not error — it
    # routes the new fn to the allocation bomb, which burns 512 MiB and spins until the
    # reap while the parent just sees None after the full timeout. Allowlist entry and
    # dispatch branch MUST land together.
    elif fn_name == "extract_xlsx_rows":
        # `max_pages` is the single int argv slot; for a workbook it means max SHEETS.
        result = _child_extract_xlsx_rows(data, max_pages)
    elif fn_name == "_test_echo":
        result = _child_test_echo(data)
    elif fn_name == "_test_crash":  # pragma: no cover — child exits nonzero
        raise RuntimeError("deliberate child crash (_test_crash, test-support)")
    elif fn_name == "_test_spin":  # pragma: no cover — child never returns
        _child_test_spin()
    else:  # pragma: no cover — _test_alloc, the only remaining allowed name
        _child_test_alloc()
    sys.stdout.write(json.dumps(result, separators=(",", ":")))
    return 0


if __name__ == "__main__":  # pragma: no cover — exercised via run_sandboxed
    sys.exit(_child_main(sys.argv[1:]))
